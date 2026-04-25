from __future__ import annotations

import csv
from io import BytesIO
from itertools import islice

from django.db.models import DecimalField, F, Q, Sum
from django.urls import reverse

from listings.models import Listing, OrderItem


REQUIRED_IMPORT_FIELDS = ("title", "price", "category")
IMAGE_FIELD_ALIASES = {
    "image_url",
    "image_urls",
    "images",
    "main_image",
    "image",
    "photo",
    "photo_url",
    "primary_image",
}


def has_seller_ai_access(user, store=None) -> bool:
    from .utils.plan_permissions import PlanPermissions

    return PlanPermissions.has_feature_access(user, "seller_ai_assistant", store=store)


def has_seller_ai_actions_access(user, store=None) -> bool:
    from .utils.plan_permissions import PlanPermissions

    return PlanPermissions.has_feature_access(user, "seller_ai_actions", store=store)


def _safe_ratio(numerator: int | float, denominator: int | float) -> int:
    if not denominator:
        return 0
    return max(0, min(100, round((numerator / denominator) * 100)))


def _normalize_column_name(name: str) -> str:
    return "".join(ch for ch in str(name or "").strip().lower() if ch.isalnum())


def _guess_import_field(column_name: str) -> str:
    normalized = _normalize_column_name(column_name)
    aliases = {
        "title": {"title", "name", "productname", "itemname", "producttitle"},
        "description": {"description", "details", "productdescription", "body"},
        "price": {"price", "amount", "cost", "sellingprice", "unitprice"},
        "stock": {"stock", "qty", "quantity", "inventory", "available"},
        "category": {"category", "productcategory", "group", "department"},
        "location": {"location", "town", "city", "area"},
        "condition": {"condition", "state", "quality"},
        "delivery_option": {"deliveryoption", "delivery", "shipping", "fulfilment", "fulfillment"},
        "brand": {"brand", "maker", "manufacturer"},
        "model": {"model", "modelnumber", "variantmodel"},
        "color": {"color", "colour"},
        "material": {"material", "fabric"},
        "image_url": {"imageurl", "image", "mainimage", "primaryimage", "photourl", "photo"},
        "image_urls": {"imageurls", "images", "galleryimages", "additionalimages"},
    }
    for field_name, options in aliases.items():
        if normalized in options:
            return field_name
    return ""


<<<<<<< HEAD
def _coerce_table(uploaded_file) -> tuple[list[str], list[list[str]]]:
=======
def _preview_records(df: pd.DataFrame, limit: int = 5) -> list[dict]:
    safe = df.head(limit).fillna("")
    records = safe.to_dict(orient="records")
    return [{str(k): str(v)[:120] for k, v in row.items()} for row in records]


def _detect_encoding(raw_bytes: bytes) -> str:
    """Detect the character encoding of raw bytes.

    Tries chardet first for accurate detection, then falls back to probing
    common encodings in order of prevalence so that CSV files exported from
    Excel (Latin-1 / Windows-1252) are handled gracefully even when chardet
    is not installed.
    """
    try:
        import chardet

        result = chardet.detect(raw_bytes)
        encoding = (result.get("encoding") or "").strip()
        if encoding:
            return encoding
    except ImportError:
        pass

    # Fallback: probe common encodings in order of prevalence.
    for candidate in ("utf-8-sig", "utf-8", "latin-1", "windows-1252", "cp1250"):
        try:
            raw_bytes.decode(candidate)
            return candidate
        except (UnicodeDecodeError, LookupError):
            continue

    # Last resort — latin-1 accepts every byte value, so it never raises.
    return "latin-1"


def _coerce_frame(uploaded_file) -> pd.DataFrame:
>>>>>>> 61f156b2678c76cd8bc4c2a5c35630d38e2d436a
    filename = (uploaded_file.name or "").lower()
    raw_bytes = uploaded_file.read()
    uploaded_file.seek(0)
    if filename.endswith(".csv"):
<<<<<<< HEAD
        decoded = raw_bytes.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(decoded.splitlines()))
        if not rows:
            return [], []
        headers = [str(cell).strip() for cell in rows[0]]
        data_rows = [[str(cell).strip() for cell in row] for row in rows[1:]]
        return headers, data_rows
=======
        encoding = _detect_encoding(raw_bytes)
        return pd.read_csv(BytesIO(raw_bytes), encoding=encoding)
>>>>>>> 61f156b2678c76cd8bc4c2a5c35630d38e2d436a
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("Excel analysis requires openpyxl to be installed.") from exc
        workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        try:
            headers = [str(cell or "").strip() for cell in next(row_iter)]
        except StopIteration:
            return [], []
        data_rows = []
        for row in row_iter:
            data_rows.append([str(cell or "").strip() for cell in row])
        return headers, data_rows
    raise ValueError("Unsupported file format. Please upload CSV or Excel.")


def run_bulk_import_preflight(uploaded_file) -> dict:
    columns, rows = _coerce_table(uploaded_file)
    if not columns and not rows:
        return {
            "summary": "Your file is empty.",
            "columns": [],
            "preview": [],
            "warnings": ["No rows were found in the uploaded file."],
            "suggestions": ["Add at least one product row before importing."],
            "field_mapping": {},
            "stats": {"rows": 0, "columns": 0},
        }

    guessed_mapping = {column: _guess_import_field(column) for column in columns}
    mapped_fields = {value for value in guessed_mapping.values() if value}
    missing_required = [field for field in REQUIRED_IMPORT_FIELDS if field not in mapped_fields]

    warnings: list[str] = []
    suggestions: list[str] = []

    title_col = next((c for c, field in guessed_mapping.items() if field == "title"), None)
    price_col = next((c for c, field in guessed_mapping.items() if field == "price"), None)
    stock_col = next((c for c, field in guessed_mapping.items() if field == "stock"), None)
    image_cols = [c for c, field in guessed_mapping.items() if field in {"image_url", "image_urls"} or _normalize_column_name(c) in IMAGE_FIELD_ALIASES]
    col_index = {column: index for index, column in enumerate(columns)}

    if missing_required:
        warnings.append(
            "Missing recommended import fields: " + ", ".join(field.replace("_", " ") for field in missing_required) + "."
        )
    if not image_cols:
        warnings.append("No image column was detected. Listings may import without product photos.")
        suggestions.append("Add an `image_url` column for the main image or `image_urls` for gallery images.")

    if title_col and title_col in col_index:
        title_idx = col_index[title_col]
        blank_titles = sum(1 for row in rows if title_idx >= len(row) or not str(row[title_idx]).strip())
        if blank_titles:
            warnings.append(f"{blank_titles} row(s) have blank titles and are likely to fail import.")

    if price_col and price_col in col_index:
        price_idx = col_index[price_col]
        invalid_prices = 0
        for row in rows:
            if price_idx >= len(row):
                continue
            raw_price = str(row[price_idx]).strip()
            cleaned = "".join(ch for ch in raw_price if ch.isdigit() or ch in ".-")
            if raw_price:
                if not cleaned:
                    invalid_prices += 1
                    continue
                try:
                    float(cleaned)
                except Exception:
                    invalid_prices += 1
        if invalid_prices:
            warnings.append(f"{invalid_prices} row(s) have prices that do not look numeric.")
            suggestions.append("Remove currency symbols or extra text from price cells before importing.")

    if stock_col and stock_col in col_index:
        stock_idx = col_index[stock_col]
        invalid_stock = 0
        for row in rows:
            if stock_idx >= len(row):
                continue
            raw_stock = str(row[stock_idx]).strip()
            cleaned = "".join(ch for ch in raw_stock if ch.isdigit() or ch == "-")
            if raw_stock:
                if not cleaned:
                    invalid_stock += 1
                    continue
                try:
                    int(cleaned)
                except Exception:
                    invalid_stock += 1
        if invalid_stock:
            warnings.append(f"{invalid_stock} row(s) have stock values that do not look numeric.")

    image_coverage = 0
    if image_cols:
        image_indexes = [col_index[column] for column in image_cols if column in col_index]
        for row in rows:
            has_image = any(idx < len(row) and str(row[idx]).strip() for idx in image_indexes)
            if has_image:
                image_coverage += 1
        if image_coverage < len(rows):
            suggestions.append(
                f"Only {image_coverage} of {len(rows)} row(s) appear to contain images. Consider filling missing image URLs."
            )

    confidence_score = 100
    confidence_score -= min(45, len(missing_required) * 15)
    confidence_score -= min(25, len(warnings) * 5)
    confidence_score = max(confidence_score, 35 if len(rows) else 0)

    summary = (
        f"Preflight reviewed {len(rows)} row(s) across {len(columns)} column(s). "
        f"Import confidence is {confidence_score}% based on required fields, images, and data cleanliness."
    )
    if confidence_score >= 85:
        suggestions.insert(0, "This file looks healthy. Proceed after confirming the field mapping.")
    elif confidence_score >= 65:
        suggestions.insert(0, "This file is workable, but you should address the highlighted warnings before import.")
    else:
        suggestions.insert(0, "This file needs cleanup before import to avoid preventable failures.")

    return {
        "summary": summary,
        "columns": columns,
        "preview": [
            {columns[idx] if idx < len(columns) else f"Column {idx + 1}": str(value)[:120] for idx, value in enumerate(row[: len(columns)])}
            for row in islice(rows, 0, 5)
        ],
        "warnings": warnings,
        "suggestions": suggestions,
        "field_mapping": guessed_mapping,
        "stats": {
            "rows": int(len(rows)),
            "columns": int(len(columns)),
            "image_coverage": image_coverage,
            "confidence_score": confidence_score,
        },
    }


def build_seller_copilot_context(user, store=None) -> dict | None:
    if not has_seller_ai_access(user, store=store):
        return None

    listing_filter = Q(seller=user) | Q(store__owner=user)
    if store is not None:
        listing_filter = Q(store=store)

    listings_qs = Listing.objects.filter(listing_filter).select_related("store", "category")
    total_listings = listings_qs.count()
    active_listings = listings_qs.filter(is_active=True, is_sold=False).count()
    low_stock = listings_qs.filter(stock__lte=5, stock__gt=0, is_active=True).count()
    out_of_stock = listings_qs.filter(stock=0, is_active=True).count()
    missing_descriptions = listings_qs.filter(Q(description__isnull=True) | Q(description="")).count()
    missing_images = listings_qs.filter(Q(image__isnull=True) | Q(image="")).count()

    order_items = OrderItem.objects.filter(listing__in=listings_qs)
    revenue = order_items.aggregate(
        total=Sum(F("price") * F("quantity"), output_field=DecimalField(max_digits=14, decimal_places=2))
    ).get("total") or 0
    recent_orders = order_items.values("order_id").distinct().count()

    priorities: list[dict] = []
    if low_stock:
        priorities.append({
            "tone": "warning",
            "title": f"{low_stock} listing(s) are running low on stock",
            "detail": "Restock or reduce promotion pressure before you lose active demand.",
        })
    if out_of_stock:
        priorities.append({
            "tone": "danger",
            "title": f"{out_of_stock} listing(s) are already out of stock",
            "detail": "Consider pausing ads or restocking them first.",
        })
    if missing_images:
        priorities.append({
            "tone": "info",
            "title": f"{missing_images} listing(s) still have weak image coverage",
            "detail": "Use bulk import with image URLs or refresh key product photos.",
        })
    if missing_descriptions:
        priorities.append({
            "tone": "muted",
            "title": f"{missing_descriptions} listing(s) need stronger descriptions",
            "detail": "The listing copilot can help tighten copy for better conversion.",
        })
    if not priorities:
        priorities.append({
            "tone": "success",
            "title": "Your catalog looks healthy right now",
            "detail": "Focus on pricing, restocking, and fresh listings to keep momentum up.",
        })

    actions = []
    try:
        actions.append({
            "label": "Open AI-ready listing form",
            "reason": "Create a better listing faster with AI assistance.",
            "url": reverse("listing-create"),
        })
    except Exception:
        pass
    if store is not None:
        try:
            actions.append({
                "label": "Run import preflight",
                "reason": "Check bulk upload data before you start an import.",
                "url": reverse("storefront:bulk_import_data", kwargs={"slug": store.slug}) + "#ai-preflight",
            })
        except Exception:
            pass
        try:
            actions.append({
                "label": "Review inventory pressure",
                "reason": "See low-stock and out-of-stock items in one place.",
                "url": reverse("storefront:inventory_dashboard", kwargs={"slug": store.slug}) + "#ai-copilot",
            })
        except Exception:
            pass

    health_score = max(
        0,
        100
        - min(35, low_stock * 4)
        - min(30, out_of_stock * 8)
        - min(20, missing_images * 3)
        - min(15, missing_descriptions * 2),
    )

    summary = (
        f"Baysoko AI Copilot reviewed {total_listings} listing(s) and {recent_orders} order-linked sale event(s). "
        f"Current revenue on record is KSh {revenue}. "
        f"The main focus areas are stock pressure, content quality, and image coverage."
    )

    return {
        "summary": summary,
        "health_score": int(health_score),
        "metrics": [
            {"label": "Active listings", "value": active_listings},
            {"label": "Low stock", "value": low_stock},
            {"label": "Out of stock", "value": out_of_stock},
            {"label": "Revenue", "value": f"KSh {revenue}"},
        ],
        "priorities": priorities[:4],
        "actions": actions[:4],
        "actions_enabled": has_seller_ai_actions_access(user, store=store),
    }
